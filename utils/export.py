"""
Exportacion a PDF y Excel.
"""

import io
import pandas as pd
from fpdf import FPDF
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows


class PDFReport(FPDF):
    def header(self):
        self.set_font("Helvetica", "B", 14)
        self.cell(0, 10, "Simulador de Precios de Casas - Reporte", align="C", new_x="LMARGIN", new_y="NEXT")
        self.ln(5)

    def footer(self):
        self.set_y(-15)
        self.set_font("Helvetica", "I", 8)
        self.cell(0, 10, f"Pagina {self.page_no()}/{{nb}}", align="C")


def exportar_pdf(resumenes, nombres, valor_uf):
    """Genera PDF con resumen comparativo y tablas."""
    pdf = PDFReport()
    pdf.alias_nb_pages()
    pdf.add_page()
    pdf.set_auto_page_break(auto=True, margin=15)

    # Valor UF
    pdf.set_font("Helvetica", "", 10)
    pdf.cell(0, 8, f"Valor UF: ${valor_uf:,.0f} CLP", new_x="LMARGIN", new_y="NEXT")
    pdf.ln(5)

    # Resumen comparativo
    pdf.set_font("Helvetica", "B", 12)
    pdf.cell(0, 10, "Resumen Comparativo", new_x="LMARGIN", new_y="NEXT")

    for i, (r, nombre) in enumerate(zip(resumenes, nombres)):
        pdf.set_font("Helvetica", "B", 10)
        pdf.set_fill_color(230, 240, 250)
        pdf.cell(0, 8, f"  {nombre}", fill=True, new_x="LMARGIN", new_y="NEXT")
        pdf.set_font("Helvetica", "", 9)

        datos = [
            ("Precio", f'{r["precio_uf"]:,.2f} UF'),
            ("Pie", f'{r["pie_monto_uf"]:,.2f} UF ({r["pie_pct"]}%)'),
            ("Monto Credito", f'{r["monto_credito_uf"]:,.2f} UF'),
            ("Tasa Efectiva", f'{r["tasa_efectiva"]}%'),
            ("Plazo", f'{r["plazo_anos"]} anos'),
            ("Sistema", r["sistema"].capitalize()),
            ("Dividendo Mes 1", f'{r["dividendo_mes1_uf"]:.2f} UF (${r["dividendo_mes1_clp"]:,.0f})'),
            ("CAE", f'{r["cae"]}%'),
            ("Total Intereses", f'{r["total_intereses_uf"]:,.2f} UF'),
            ("Costo Total", f'{r["costo_total_uf"]:,.2f} UF (${r["costo_total_clp"]:,.0f})'),
            ("Valor Futuro", f'{r["valor_futuro_final_uf"]:,.2f} UF'),
            ("Ganancia Neta", f'{r["ganancia_neta_uf"]:,.2f} UF'),
            ("Arriendo Estimado", f'{r["arriendo_mensual_uf"]:.2f} UF/mes'),
            ("Flujo Neto Promedio", f'{r["flujo_neto_promedio_uf"]:.2f} UF/mes'),
        ]

        for label, valor in datos:
            pdf.cell(60, 6, f"    {label}:", new_x="RIGHT")
            pdf.cell(0, 6, valor, new_x="LMARGIN", new_y="NEXT")

        pdf.ln(3)

    # Tabla de amortizacion resumida (anual) por escenario
    for i, (r, nombre) in enumerate(zip(resumenes, nombres)):
        pdf.add_page()
        pdf.set_font("Helvetica", "B", 11)
        pdf.cell(0, 8, f"Tabla de Amortizacion Anual - {nombre}", new_x="LMARGIN", new_y="NEXT")
        pdf.ln(3)

        tabla = r["tabla_amortizacion"]
        if tabla.empty:
            continue

        anual = tabla.groupby("ano").agg({
            "cuota_base": "sum",
            "interes": "sum",
            "capital": "sum",
            "seg_desgravamen": "sum",
            "seg_incendio": "sum",
            "cuota_total": "sum",
        }).reset_index()
        anual["saldo"] = tabla.groupby("ano")["saldo"].last().values

        # Encabezados
        pdf.set_font("Helvetica", "B", 7)
        pdf.set_fill_color(200, 200, 200)
        cols = ["Ano", "Cuota", "Interes", "Capital", "Saldo", "Seguros", "Total"]
        widths = [15, 25, 25, 25, 30, 25, 30]
        for col, w in zip(cols, widths):
            pdf.cell(w, 6, col, border=1, fill=True, align="C")
        pdf.ln()

        # Filas
        pdf.set_font("Helvetica", "", 7)
        for _, row in anual.iterrows():
            vals = [
                f'{int(row["ano"])}',
                f'{row["cuota_base"]:,.2f}',
                f'{row["interes"]:,.2f}',
                f'{row["capital"]:,.2f}',
                f'{row["saldo"]:,.2f}',
                f'{row["seg_desgravamen"] + row["seg_incendio"]:,.2f}',
                f'{row["cuota_total"]:,.2f}',
            ]
            for val, w in zip(vals, widths):
                pdf.cell(w, 5, val, border=1, align="R")
            pdf.ln()

    buffer = io.BytesIO()
    pdf.output(buffer)
    buffer.seek(0)
    return buffer


def exportar_excel(resumenes, nombres, valor_uf):
    """Genera Excel con hojas por escenario."""
    buffer = io.BytesIO()
    wb = Workbook()

    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font_white = Font(bold=True, color="FFFFFF", size=10)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # Hoja resumen
    ws_resumen = wb.active
    ws_resumen.title = "Resumen"

    ws_resumen.cell(row=1, column=1, value="Simulador de Precios de Casas").font = Font(bold=True, size=14)
    ws_resumen.cell(row=2, column=1, value=f"Valor UF: ${valor_uf:,.0f} CLP")

    encabezados = [
        "Escenario", "Precio (UF)", "Pie (%)", "Pie (UF)", "Monto Credito (UF)",
        "Tasa Efectiva (%)", "Plazo (anos)", "Sistema", "Dividendo Mes 1 (UF)",
        "Dividendo Mes 1 ($)", "CAE (%)", "Total Intereses (UF)", "Costo Total (UF)",
        "Costo Total ($)", "Valor Futuro (UF)", "Ganancia Neta (UF)",
        "Arriendo (UF/mes)", "Flujo Neto Prom (UF/mes)",
    ]

    for col, enc in enumerate(encabezados, 1):
        cell = ws_resumen.cell(row=4, column=col, value=enc)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    for row_idx, (r, nombre) in enumerate(zip(resumenes, nombres), 5):
        datos = [
            nombre, r["precio_uf"], r["pie_pct"], r["pie_monto_uf"],
            r["monto_credito_uf"], r["tasa_efectiva"], r["plazo_anos"],
            r["sistema"].capitalize(), r["dividendo_mes1_uf"],
            r["dividendo_mes1_clp"], r["cae"], r["total_intereses_uf"],
            r["costo_total_uf"], r["costo_total_clp"],
            r["valor_futuro_final_uf"], r["ganancia_neta_uf"],
            r["arriendo_mensual_uf"], r["flujo_neto_promedio_uf"],
        ]
        for col, val in enumerate(datos, 1):
            cell = ws_resumen.cell(row=row_idx, column=col, value=val)
            cell.border = thin_border

    # Ajustar anchos
    for col in range(1, len(encabezados) + 1):
        ws_resumen.column_dimensions[chr(64 + min(col, 26))].width = 18

    # Hojas por escenario con tabla de amortizacion
    for r, nombre in zip(resumenes, nombres):
        safe_name = nombre[:28].replace("/", "-")
        ws = wb.create_sheet(title=safe_name)

        ws.cell(row=1, column=1, value=f"Tabla de Amortizacion - {nombre}").font = Font(bold=True, size=12)

        tabla = r["tabla_amortizacion"]
        if tabla.empty:
            continue

        cols_tabla = ["mes", "ano", "cuota_base", "interes", "capital", "saldo",
                      "seg_desgravamen", "seg_incendio", "cuota_total"]
        headers_es = ["Mes", "Ano", "Cuota Base", "Interes", "Capital", "Saldo",
                      "Seg. Desgravamen", "Seg. Incendio", "Cuota Total"]

        for col, h in enumerate(headers_es, 1):
            cell = ws.cell(row=3, column=col, value=h)
            cell.font = header_font_white
            cell.fill = header_fill
            cell.border = thin_border

        for row_idx, (_, fila) in enumerate(tabla.iterrows(), 4):
            for col, c in enumerate(cols_tabla, 1):
                cell = ws.cell(row=row_idx, column=col, value=fila[c])
                cell.border = thin_border
                if col >= 3:
                    cell.number_format = "#,##0.00"

        # Flujo arriendo
        flujo = r["flujo_arriendo"]
        if not flujo.empty:
            ws_flujo = wb.create_sheet(title=f"{safe_name[:20]} Flujo")
            ws_flujo.cell(row=1, column=1, value=f"Flujo de Arriendo - {nombre}").font = Font(bold=True, size=12)

            cols_flujo = ["mes", "ano", "arriendo", "dividendo", "gastos_comunes", "contribuciones", "flujo_neto"]
            headers_flujo = ["Mes", "Ano", "Arriendo (UF)", "Dividendo (UF)", "Gastos Comunes",
                           "Contribuciones", "Flujo Neto (UF)"]

            for col, h in enumerate(headers_flujo, 1):
                cell = ws_flujo.cell(row=3, column=col, value=h)
                cell.font = header_font_white
                cell.fill = header_fill
                cell.border = thin_border

            for row_idx, (_, fila) in enumerate(flujo.iterrows(), 4):
                for col, c in enumerate(cols_flujo, 1):
                    cell = ws_flujo.cell(row=row_idx, column=col, value=fila[c])
                    cell.border = thin_border
                    if col >= 3:
                        cell.number_format = "#,##0.0000"

    wb.save(buffer)
    buffer.seek(0)
    return buffer
